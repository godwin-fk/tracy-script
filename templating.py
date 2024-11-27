import pandas as pd
from openpyxl import load_workbook

def replace_master_data(file_path, df, output_file_path=None):
    """
    Replace all data in the 'Master Data' sheet with a pandas DataFrame, 
    while keeping other sheets intact, and place 'Master Data' as the first tab.

    :param file_path: Path to the existing Excel file.
    :param df: Pandas DataFrame containing the new data to replace in the 'Master Data' sheet.
    :param output_file_path: Path to save the updated file. If None, overwrites the original file.
    """
    try:
        save_path = output_file_path if output_file_path else file_path
        
        # Load the existing workbook
        workbook = load_workbook(file_path)
        
        # Check if 'master data' sheet exists and remove it
        if 'master data' in workbook.sheetnames:
            del workbook['master data']
        
        # Create a new sheet for 'master data'
        master_data_sheet = workbook.create_sheet('master data')
        
        # Write DataFrame content to the new sheet
        for row_idx, row_data in enumerate(df.itertuples(index=False, name=None), start=1):
            for col_idx, value in enumerate(row_data, start=1):
                master_data_sheet.cell(row=row_idx + 1, column=col_idx, value=value)
        
        # Add headers
        for col_idx, column_name in enumerate(df.columns, start=1):
            master_data_sheet.cell(row=1, column=col_idx, value=column_name)
        
        # Reorder sheets to make 'master data' the first sheet
        workbook._sheets.insert(0, workbook._sheets.pop(workbook._sheets.index(master_data_sheet)))
        
        # Save the workbook
        workbook.save(save_path)
        
        print(f"Data in 'Master Data' replaced successfully in '{save_path}'.")
    
    except Exception as e:
        print(f"An error occurred: {e}")

# Example usage
file_path = "./report-template/notifier-template.xlsx"  # Input Excel file
output_file_path = None  # Optional: Path to save the updated file (None means overwrite the original)

# Create DataFrame from a CSV file
csv_file_path = "./dist/notifer-report-name"  # change to the correct notifier report name
df = pd.read_csv(csv_file_path)

replace_master_data(file_path, df, output_file_path)
