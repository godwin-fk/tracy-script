import os
import logging
import base64
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from datetime import datetime, timedelta, timezone
from openpyxl import load_workbook, Workbook
import pandas as pd

logger = logging.getLogger(__name__)

def get_unique_filepath(filepath):
    base, extension = os.path.splitext(filepath)
    counter = 1
    while os.path.exists(filepath):
        filepath = f"{base}_{counter}{extension}"
        counter += 1
    return filepath

def get_gmail_service(shipper_id: str, agent_id: str):
    """Authenticate using environment variables and return the Gmail API service."""
    shipper_id = shipper_id.replace("-", "_")
    token_info = {
            "client_id": os.getenv(f"{shipper_id}_{agent_id}_GOOGLE_CLIENT_ID"),
            "client_secret": os.getenv(f"{shipper_id}_{agent_id}_GOOGLE_CLIENT_SECRET"),
            "refresh_token": os.getenv(f"{shipper_id}_{agent_id}_GOOGLE_REFRESH_TOKEN"),
            "token_uri": os.getenv(f"GOOGLE_TOKEN_URI"),
            "scopes": ["https://mail.google.com/"]
    }
    for key, value in token_info.items():
        if not value:
            raise ValueError(f"Missing required environment variable: {key}")

    try:
        creds = Credentials(
            None,
            refresh_token=token_info["refresh_token"],
            client_id=token_info["client_id"],
            client_secret=token_info["client_secret"],
            token_uri=token_info["token_uri"],
        )

        return build("gmail", "v1", credentials=creds)
    except Exception as e:
        logger.error(f"Error authenticating with Gmail API: {e}")
        raise

def search_emails_with_attachments(gmail_service, query, save_path):
    """Search emails matching the query and download attachments to the save path."""
    os.makedirs(save_path, exist_ok=True)

    try:
        # Dict to store received date along with filePath
        received_dates = {}
        # Search emails matching the query
        results = gmail_service.users().messages().list(userId="me", q=query).execute()
        messages = results.get("messages", [])
        

        if not messages:
            print("No emails found matching the query.")
            return

        for message in messages:
            msg = gmail_service.users().messages().get(userId="me", id=message["id"]).execute()
            received_timestamp = int(msg.get("internalDate")) / 1000
            received_date = datetime.fromtimestamp(received_timestamp, tz=timezone.utc).strftime("%d:%m:%Y")

            print(received_date)
            subject = next(
                (header["value"] for header in msg["payload"]["headers"] if header["name"] == "Subject"), "No Subject"
            )
            print(f"Found Email - Subject: {subject}")

            # Download attachments
            if "parts" in msg["payload"]:
                for part in msg["payload"]["parts"]:
                    if part.get("filename") and part["body"].get("attachmentId"):
                        attachment_id = part["body"]["attachmentId"]
                        attachment = gmail_service.users().messages().attachments().get(
                            userId="me", messageId=message["id"], id=attachment_id
                        ).execute()
                        data = attachment["data"]
                        filepath = get_unique_filepath(os.path.join(save_path, part["filename"]))
                        received_dates[filepath] = received_date
                        with open(filepath, "wb") as f:
                            f.write(base64.urlsafe_b64decode(data))
                        print(f"Downloaded Attachment: {part['filename']}")
        # print(received_date_filePath)
        return received_dates
    except HttpError as error:
        print(f"An error occurred: {error}")
        
def merge_xlsx_files(directory, received_dates, output_file="./temp/merged_holdover_reports.xlsx"):
        temp_directory = os.path.dirname(output_file)
        os.makedirs(temp_directory, exist_ok=True)
        """Merge all Excel files in a directory into a single workbook."""
        
        merged_workbook = Workbook()
        merged_sheet = merged_workbook.active
        merged_sheet.title = "Merged Data"
        headers_written = False
        
        headers = [
            "Load Number", "CARRIER", "CONTAINER ID", "DESTINATION CITY", 
            "DESTINATION STATE", "DD DATE", "DD TIME", "BILL DATE", 
            "BILL TIME", "ZDLT LATE CODE", "ON TIME? (Y/N)", "SPLIT? (Y/N)", 
            "FRESH PRIORITY STO? (Y/N)", "NOTES/COMMENTS"
        ]

        merged_data = pd.DataFrame()
        for i,file_name in enumerate(os.listdir(directory)):
            if file_name.endswith(".xlsx"):
                file_path = os.path.join(directory, file_name)
    
                df = pd.read_excel(file_path, sheet_name=0,header=None)
                plant_row_index = df[df.eq("PLANT:").any(axis=1)].index[0]
                plant_column_index = df.loc[plant_row_index].eq("PLANT:").idxmax()
                result_column_index = df.columns.get_loc(plant_column_index) + 1
                FACILITY = df.iloc[plant_row_index, result_column_index]
                
                header_row_index = df[df.eq("SHIPMENT NUMBER").any(axis=1)].index[0]
                # print(header_row_index)
                data = df.iloc[header_row_index+1:].reset_index(drop=True)
                data.columns = headers
                
                # data.columns = df.iloc[header_row_index]  # Set the proper headers
                
                # print(data)
                # data.columns = [str(col).strip() for col in data.columns]

                data['PLANT'] = FACILITY 
                if file_path in received_dates:
                    data['RECEIVED DATE'] = received_dates[file_path]    
                    
                # data.rename( columns={'Unnamed: 14':'PLANT'}, inplace=True )
                # data.rename( columns={'Unnamed: 15':'RECEIVED DATE'}, inplace=True )

                merged_data = pd.concat([merged_data, data], ignore_index=True)
        merged_data.to_excel(output_file, index=False)
        print(f"Merged file saved as: {output_file}")
        return output_file


# def merge_xlsx_files(directory, output_file="../temp/merged_holdover_reports.xlsx"):
#     # Create a new workbook for the merged file
#     merged_workbook = Workbook()
#     merged_sheet = merged_workbook.active
#     merged_sheet.title = "Merged Data"
    
#     # Flag to track if headers have been written to the merged file
#     headers_written = False

#     # Iterate over all files in the directory
#     for file_name in os.listdir(directory):
#         if file_name.endswith(".xlsx"):
#             file_path = os.path.join(directory, file_name)
#             print(f"Processing: {file_path}")
            
#             # Open the workbook
#             workbook = load_workbook(file_path)
#             sheet = workbook.active  # Assuming data is in the first sheet

#             # Read data from the current workbook
#             for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
#                 # Write headers only once
#                 if row_index == 1 and headers_written:
#                     continue
#                 merged_sheet.append(row)
            
#             headers_written = True  # Headers have now been written
    
#     # Save the merged workbook
#     merged_workbook.save(output_file)
#     print(f"Merged file saved as: {output_file}")
#     return output_file


def clean_excel(file_path: str, output_path: str):
    # Read the Excel file
    df = pd.read_excel(file_path)

    # Get the name of the first column
    first_col = df.columns[0]

    def is_convertible(value):
        """
        Checks if the value can be converted to an integer.
        Returns True if convertible, False otherwise.
        """
        try:
            int(value)
            return True
        except (ValueError, TypeError):
            return False

    # Filter rows where the first column value is convertible to int
    cleaned_df = df[df[first_col].apply(is_convertible)]

    # Save the cleaned DataFrame to a new Excel file
    cleaned_df.to_excel(output_path, index=False)
    print(f"Cleaned Excel file saved to: {output_path}")



def update_headers(output_file: str):
    """
    Removes the first row of the given Excel file and adds specified headers.
    
    Parameters:
    - output_file: str - Path to the Excel file to update.
    """
    # Define the new headers
    headers = [
        "Load Number", "CARRIER", "CONTAINER ID", "DESTINATION CITY", 
        "DESTINATION STATE", "DD DATE", "DD TIME", "BILL DATE", 
        "BILL TIME", "ZDLT LATE CODE", "ON TIME? (Y/N)", "SPLIT? (Y/N)", 
        "FRESH PRIORITY STO? (Y/N)", "NOTES/COMMENTS"
    ]
    
    # Load the workbook
    wb = load_workbook(output_file)
    sheet = wb.active  # Assume we're working with the first sheet
    
    # Remove the first row
    sheet.delete_rows(1)
    
    # Insert new headers at the first row
    sheet.insert_rows(1)
    for col_index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_index).value = header
    
    wb.save(output_file)
    print(f"Headers updated and first row removed in: {output_file}")


def convert_excel_to_csv(output_file: str, shipper_id_holdover: str, start_date: str, end_date: str):
    csv_filepath = f"../temp/{shipper_id_holdover}-{start_date}_{end_date}.csv"
    
    df = pd.read_excel(output_file)
    
    df.to_csv(csv_filepath, index=False)
    print(f"Converted Excel to CSV and saved as: {csv_filepath}")



if __name__ == "__main__":
    
    # Replace these with appropriate values
    shipper_id = "smithfield_foods"
    agent_id = "TRACY"
    save_path = "attachments"
    output_file = "../temp/output_data.xlsx"

    start_date = "2024-12-05"
    end_date = "2024-12-05"

    start_timestamp = int(datetime.strptime(start_date, "%Y-%m-%d").timestamp())
    
    # Add one day to end_date to make it inclusive
    end_date_inclusive = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
    end_timestamp = int(end_date_inclusive.timestamp())

    search_query = f'holdover OR "hold over" has:attachment in:anywhere after:{start_timestamp} before:{end_timestamp}'

    gmail_service = get_gmail_service(shipper_id, agent_id)
    received_dates = search_emails_with_attachments(gmail_service, search_query, save_path)

    input_file = merge_xlsx_files(save_path,received_dates)
    clean_excel(input_file, output_file)
    update_headers(output_file)
    
    shipper_id_holdover = f'{shipper_id}-holdover'
    convert_excel_to_csv(output_file, shipper_id_holdover, start_date, end_date)