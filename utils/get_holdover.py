import os
import logging
import base64
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from datetime import datetime, timedelta, timezone
from openpyxl import load_workbook, Workbook
import pandas as pd

logger = logging.getLogger(__name__)


class GmailDataProcessor:
    def __init__(self, shipper_id, agent_id):
        self.shipper_id = shipper_id.replace("-", "_")
        self.agent_id = agent_id
        self.gmail_service = self.authenticate_gmail_service()
    
    def authenticate_gmail_service(self):
        """Authenticate using environment variables and return the Gmail API service."""
        shipper_id = self.shipper_id.replace("-", "_")
        token_info = {
                "client_id": os.getenv(f"{self.shipper_id}_{self.agent_id}_GOOGLE_CLIENT_ID"),
                "client_secret": os.getenv(f"{self.shipper_id}_{self.agent_id}_GOOGLE_CLIENT_SECRET"),
                "refresh_token": os.getenv(f"{self.shipper_id}_{self.agent_id}_GOOGLE_REFRESH_TOKEN"),
                "token_uri": os.getenv(f"GOOGLE_TOKEN_URI"),
                "scopes": ["https://mail.google.com/"]
        }
        for key, value in token_info.items():
            if not value:
                raise ValueError(f"Missing required environment variable: {key}")

        try:
            creds = Credentials(
                None,
                refresh_token=token_info["refresh_token"],
                client_id=token_info["client_id"],
                client_secret=token_info["client_secret"],
                token_uri=token_info["token_uri"],
            )

            return build("gmail", "v1", credentials=creds)
        except Exception as e:
            logger.error(f"Error authenticating with Gmail API: {e}")
            raise

    @staticmethod
    def get_unique_filepath(filepath):
        """Generate a unique file path by appending a counter if the file exists."""
        base, extension = os.path.splitext(filepath)
        counter = 1
        while os.path.exists(filepath):
            filepath = f"{base}_{counter}{extension}"
            counter += 1
        return filepath

    def search_emails_with_attachments(self, query, save_path):
        """Search emails matching the query and download attachments to the save path."""
        os.makedirs(save_path, exist_ok=True)

        try:
            # Dict to store received date along with filePath
            received_dates = {}
            # Search emails matching the query
            results = self.gmail_service.users().messages().list(userId="me", q=query).execute()
            messages = results.get("messages", [])
            

            if not messages:
                print("No emails found matching the query.")
                return

            for message in messages:
                msg = self.gmail_service.users().messages().get(userId="me", id=message["id"]).execute()
                received_timestamp = int(msg.get("internalDate")) / 1000
                received_date = datetime.fromtimestamp(received_timestamp, tz=timezone.utc).strftime("%d:%m:%Y")

                print(received_date)
                subject = next(
                    (header["value"] for header in msg["payload"]["headers"] if header["name"] == "Subject"), "No Subject"
                )
                print(f"Found Email - Subject: {subject}")

                # Download attachments
                if "parts" in msg["payload"]:
                    for part in msg["payload"]["parts"]:
                        if part.get("filename") and part["body"].get("attachmentId"):
                            attachment_id = part["body"]["attachmentId"]
                            attachment = self.gmail_service.users().messages().attachments().get(
                                userId="me", messageId=message["id"], id=attachment_id
                            ).execute()
                            data = attachment["data"]
                            filepath = self.get_unique_filepath(os.path.join(save_path, part["filename"]))
                            received_dates[filepath] = received_date
                            with open(filepath, "wb") as f:
                                f.write(base64.urlsafe_b64decode(data))
                            print(f"Downloaded Attachment: {part['filename']}")
            # print(received_date_filePath)
            return received_dates
        except HttpError as error:
            print(f"An error occurred: {error}")

    @ staticmethod
    def merge_xlsx_files(directory, received_dates):
        """Merge all Excel files in a directory into a single workbook."""
        output_file = os.path.join('./tmp', "merged_holdover_reports.xlsx")

        merged_workbook = Workbook()
        merged_sheet = merged_workbook.active
        merged_sheet.title = "Merged Data"
        headers_written = False
        
        headers = [
            "Load Number", "CARRIER", "CONTAINER ID", "DESTINATION CITY", 
            "DESTINATION STATE", "DD DATE", "DD TIME", "BILL DATE", 
            "BILL TIME", "ZDLT LATE CODE", "ON TIME? (Y/N)", "SPLIT? (Y/N)", 
            "FRESH PRIORITY STO? (Y/N)", "NOTES/COMMENTS"
        ]

        merged_data = pd.DataFrame()
        for i,file_name in enumerate(os.listdir(directory)):
            if file_name.endswith(".xlsx"):
                file_path = os.path.join(directory, file_name)
    
                df = pd.read_excel(file_path, sheet_name=0,header=None)
                plant_row_index = df[df.eq("PLANT:").any(axis=1)].index[0]
                plant_column_index = df.loc[plant_row_index].eq("PLANT:").idxmax()
                result_column_index = df.columns.get_loc(plant_column_index) + 1
                FACILITY = df.iloc[plant_row_index, result_column_index]
                
                header_row_index = df[df.eq("SHIPMENT NUMBER").any(axis=1)].index[0]

                data = df.iloc[header_row_index+1:].reset_index(drop=True)
                data.columns = headers  
                # data.columns = df.iloc[header_row_index]  # Set the proper headers
    
                data['PLANT'] = FACILITY 
                if file_path in received_dates:
                    data['RECEIVED DATE'] = received_dates[file_path]    
                    
                # data.rename( columns={'Unnamed: 14':'PLANT'}, inplace=True )
                # data.rename( columns={'Unnamed: 15':'RECEIVED DATE'}, inplace=True )

                merged_data = pd.concat([merged_data, data], ignore_index=True)
        merged_data.to_excel(output_file, index=False)
        print(f"Merged file saved as: {output_file}")
        return output_file
    
    @staticmethod
    def clean_excel(file_path, output_path):
        """Clean an Excel file by filtering rows where the first column is convertible to int."""
        df = pd.read_excel(file_path)
        first_col = df.columns[0]

        def is_convertible(value):
            try:
                int(value)
                return True
            except (ValueError, TypeError):
                return False

        cleaned_df = df[df[first_col].apply(is_convertible)]
        cleaned_df.to_excel(output_path, index=False)
        print(f"Cleaned Excel file saved to: {output_path}")

    @staticmethod
    def update_headers(output_file):
        """Update headers of an Excel file."""
        headers = [
            "Load Number", "CARRIER", "CONTAINER ID", "DESTINATION CITY",
            "DESTINATION STATE", "DD DATE", "DD TIME", "BILL DATE",
            "BILL TIME", "ZDLT LATE CODE", "ON TIME? (Y/N)", "SPLIT? (Y/N)",
            "FRESH PRIORITY STO? (Y/N)", "NOTES/COMMENTS"
        ]

        wb = load_workbook(output_file)
        sheet = wb.active
        sheet.delete_rows(1)
        sheet.insert_rows(1)
        for col_index, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_index).value = header

        wb.save(output_file)
        print(f"Headers updated in: {output_file}")

    @staticmethod
    def convert_excel_to_csv(output_file, csv_filepath):
        """Convert an Excel file to CSV."""
        df = pd.read_excel(output_file)
        df.to_csv(csv_filepath, index=False)
        print(f"Converted Excel to CSV and saved as: {csv_filepath}")

    def process_emails(self, save_path, output_file, start_date, end_date):
        """Orchestrate the entire process: download attachments, merge, clean, and convert."""
        start_timestamp = int(datetime.strptime(start_date, "%Y-%m-%d").timestamp())
        end_date_inclusive = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
        end_timestamp = int(end_date_inclusive.timestamp())
        search_query = f'holdover OR "hold over" has:attachment in:anywhere after:{start_timestamp} before:{end_timestamp}'

        received_dates = self.search_emails_with_attachments(search_query, save_path)
        input_file = self.merge_xlsx_files(save_path, received_dates)
        self.clean_excel(input_file, output_file)
        self.update_headers(output_file)
        shipper_id_holdover = f"{self.shipper_id}-holdover"
        csv_filepath = f"./tmp/{shipper_id_holdover}-{start_date}_{end_date}.csv"
        self.convert_excel_to_csv(output_file, csv_filepath)


# if __name__ == "__main__":
#     processor = GmailDataProcessor("smithfield_foods", "TRACY")
#     output_file = "./tmp/output_data.xlsx"
#     os.makedirs(os.path.dirname(output_file), exist_ok=True)
#     processor.process_emails(
#         save_path="./tmp/attachments",
#         output_file=output_file,
#         start_date="2024-12-05",
#         end_date="2024-12-05"
#     )
